from __future__ import annotations

import csv
import io
import json
import re
import time
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup

from cortex_engine.journal_authority import classify_journal_authority
from cortex_engine.preface_classification import classify_credibility_tier_with_reason
from cortex_engine.utils.logging_utils import get_logger

logger = get_logger(__name__)

_DOI_RE = re.compile(r"10\.\d{4,9}/\S+", re.IGNORECASE)
_YEAR_RE = re.compile(r"\b(19|20)\d{2}\b")
_PMID_RE = re.compile(r"^\d{1,8}$")
_PREPRINT_MARKERS = (
    "preprint",
    "pre-print",
    "arxiv",
    "biorxiv",
    "medrxiv",
    "ssrn",
    "research square",
)
_RESEARCH_COLUMN_ALIASES = {
    "title": ["title", "article title", "paper title"],
    "authors": ["author", "authors", "author s", "author(s)"],
    "year": ["year", "published year", "pub year", "date"],
    "doi": ["doi", "digital object identifier"],
    "journal": ["journal", "source", "publication", "periodical"],
    "abstract": ["abstract", "summary"],
    "volume": ["volume", "vol"],
    "issue": ["issue", "no"],
    "pages": ["pages", "page range"],
    "accession": ["accession number", "pmid", "pubmed id"],
    "aim": ["aim", "aim objective", "objective", "purpose"],
    "notes": ["notes", "tags", "study", "ref"],
}


def _normalize_text(text: str) -> str:
    base = unicodedata.normalize("NFKD", str(text or ""))
    ascii_text = base.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_text.lower()
    cleaned = re.sub(r"[^a-z0-9]+", " ", lowered).strip()
    return re.sub(r"\s+", " ", cleaned)


def _normalize_header_key(text: str) -> str:
    return _normalize_text(str(text or "").replace("&", " and "))


def _preview_confidence(citation: Dict[str, Any]) -> str:
    if str(citation.get("doi") or "").strip():
        return "green"
    if str(citation.get("title") or "").strip() and str(citation.get("authors") or "").strip() and _extract_year(citation.get("year")):
        return "amber"
    if str(citation.get("title") or "").strip():
        return "red"
    return "invalid"


def _detect_column_mapping(headers: List[str]) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    used_fields: set[str] = set()
    normalized_headers = [_normalize_header_key(header) for header in headers]

    for idx, header in enumerate(normalized_headers):
        for canonical, aliases in _RESEARCH_COLUMN_ALIASES.items():
            if canonical in used_fields:
                continue
            if header == canonical or header in aliases:
                mapping[idx] = canonical
                used_fields.add(canonical)
                break

    for idx, header in enumerate(normalized_headers):
        if idx in mapping or not header:
            continue
        for canonical, aliases in _RESEARCH_COLUMN_ALIASES.items():
            if canonical in used_fields:
                continue
            choices = [canonical, *aliases]
            if any(header in choice or choice in header for choice in choices):
                mapping[idx] = canonical
                used_fields.add(canonical)
                break

    return mapping


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_citation_from_row(
    row_values: List[Any],
    headers: List[str],
    mapping: Dict[int, str],
    row_id: int,
) -> Optional[Dict[str, Any]]:
    values = [_stringify_cell(value) for value in row_values]
    if not any(values):
        return None

    citation = {
        "row_id": row_id,
        "title": "",
        "authors": "",
        "year": "",
        "doi": "",
        "journal": "",
        "abstract": "",
        "volume": "",
        "issue": "",
        "pages": "",
        "accession": "",
        "aim": "",
        "notes": "",
        "extra_fields": {},
    }
    extras: Dict[str, str] = {}
    for idx, value in enumerate(values):
        if not value:
            continue
        canonical = mapping.get(idx)
        if canonical:
            citation[canonical] = value
        else:
            header = str(headers[idx] or f"column_{idx + 1}").strip()
            extras[header] = value

    citation["doi"] = _normalize_doi(citation.get("doi") or "")
    citation["year"] = _extract_year(citation.get("year"))
    citation["extra_fields"] = extras
    if not str(citation.get("title") or "").strip():
        return None
    citation["preview_confidence"] = _preview_confidence(citation)
    return citation


def _parse_delimited_text(raw_text: str) -> List[List[str]]:
    text = str(raw_text or "").replace("\r\n", "\n").replace("\r", "\n").strip("\n")
    if not text.strip():
        return []
    delimiter = "\t" if "\t" in text.splitlines()[0] else ","
    if delimiter == ",":
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
            delimiter = getattr(dialect, "delimiter", ",") or ","
        except Exception:
            delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader]


def _parse_xlsx_bytes(data: bytes) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX parsing") from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows: List[List[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_stringify_cell(cell) for cell in row])
    return rows


def _build_parse_result(rows: List[List[str]], source_name: str) -> Dict[str, Any]:
    if not rows:
        return {
            "source_name": source_name,
            "citations": [],
            "detected_fields": [],
            "warnings": ["No rows detected."],
        }

    headers = [_stringify_cell(cell) or f"column_{idx + 1}" for idx, cell in enumerate(rows[0])]
    mapping = _detect_column_mapping(headers)
    citations: List[Dict[str, Any]] = []
    warnings: List[str] = []
    seen_dois: set[str] = set()
    duplicate_dois: List[str] = []

    for offset, row in enumerate(rows[1:], start=1):
        citation = _build_citation_from_row(row, headers, mapping, row_id=offset)
        if not citation:
            continue
        doi = str(citation.get("doi") or "").lower()
        if doi:
            if doi in seen_dois:
                duplicate_dois.append(citation["doi"])
                continue
            seen_dois.add(doi)
        citations.append(citation)

    if duplicate_dois:
        warnings.append(
            f"Skipped {len(duplicate_dois)} duplicate DOI row(s): {', '.join(sorted(dict.fromkeys(duplicate_dois))[:5])}"
        )
    if "title" not in mapping.values():
        warnings.append("No title column detected. Resolution requires a title column.")

    return {
        "source_name": source_name,
        "citations": citations,
        "detected_fields": [field for field in _RESEARCH_COLUMN_ALIASES.keys() if field in mapping.values()],
        "warnings": warnings,
    }


def parse_research_spreadsheet_text(raw_text: str, source_name: str = "Pasted text") -> Dict[str, Any]:
    return _build_parse_result(_parse_delimited_text(raw_text), source_name)


def parse_research_spreadsheet_upload(filename: str, data: bytes) -> Dict[str, Any]:
    suffix = Path(str(filename or "upload")).suffix.lower()
    if suffix == ".xlsx":
        rows = _parse_xlsx_bytes(data)
    else:
        rows = _parse_delimited_text(data.decode("utf-8", errors="ignore"))
    return _build_parse_result(rows, Path(str(filename or "upload")).name)


def build_research_preview_rows(citations: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for citation in citations or []:
        extra_fields = dict(citation.get("extra_fields") or {})
        rows.append(
            {
                "keep": True,
                "row_id": citation.get("row_id"),
                "title": citation.get("title", ""),
                "authors": citation.get("authors", ""),
                "year": citation.get("year", ""),
                "doi": citation.get("doi", ""),
                "journal": citation.get("journal", ""),
                "accession": citation.get("accession", ""),
                "source_table": extra_fields.get("table_number") or "",
                "source_trial": extra_fields.get("trial_label") or "",
                "source_group": extra_fields.get("combined_group") or extra_fields.get("group_label") or "",
                "source_ref": extra_fields.get("reference_number") or "",
                "confidence": citation.get("preview_confidence") or _preview_confidence(citation),
            }
        )
    return rows


def build_research_preferred_url_list(resolved: List[Dict[str, Any]]) -> List[str]:
    urls: List[str] = []
    seen: set[str] = set()
    for item in resolved or []:
        open_access = item.get("open_access") or {}
        candidate = str(open_access.get("pdf_url") or item.get("resolved_url") or "").strip()
        if candidate and candidate not in seen:
            seen.add(candidate)
            urls.append(candidate)
    return urls


def _title_similarity(left: str, right: str) -> float:
    a = _normalize_text(left)
    b = _normalize_text(right)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    ratio = SequenceMatcher(a=a, b=b).ratio()
    a_tokens = set(a.split())
    b_tokens = set(b.split())
    jaccard = (len(a_tokens & b_tokens) / len(a_tokens | b_tokens)) if (a_tokens or b_tokens) else 0.0
    return round((ratio * 0.7) + (jaccard * 0.3), 3)


def _extract_first_author(authors: str) -> str:
    text = str(authors or "").strip()
    if not text:
        return ""
    first = re.split(r";|\band\b", text, maxsplit=1, flags=re.IGNORECASE)[0].strip()
    if "," in first:
        return first.split(",", 1)[0].strip()
    parts = [part for part in re.split(r"\s+", first) if part]
    return parts[-1] if parts else ""


def _normalize_doi(raw: str) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    text = text.replace("https://doi.org/", "").replace("http://doi.org/", "")
    text = text.replace("https://dx.doi.org/", "").replace("http://dx.doi.org/", "")
    text = re.sub(r"^doi:\s*", "", text, flags=re.IGNORECASE)
    match = _DOI_RE.search(text)
    return match.group(0).rstrip(" .;,)") if match else text.strip().rstrip(" .;,)")


def _extract_year(raw: Any) -> str:
    text = str(raw or "").strip()
    match = _YEAR_RE.search(text)
    return match.group(0) if match else ""


def _pick_primary_url(message: Dict[str, Any], doi: str) -> str:
    candidates = [
        str(message.get("URL") or "").strip(),
        str(message.get("link") or "").strip(),
        f"https://doi.org/{doi}" if doi else "",
    ]
    for candidate in candidates:
        if candidate:
            return candidate
    return ""


def _empty_open_access() -> Dict[str, Any]:
    return {"is_oa": None, "oa_status": "unknown", "pdf_url": "", "oa_source": ""}


def _citation_source_metadata(citation: Dict[str, Any]) -> Dict[str, Any]:
    extra_fields = dict(citation.get("extra_fields") or {})
    source_review = str(extra_fields.get("source_review") or citation.get("source_review") or "").strip()
    source_review_title = str(extra_fields.get("source_review_title") or citation.get("source_review_title") or "").strip()
    source_doc_id = str(extra_fields.get("source_doc_id") or citation.get("source_doc_id") or "").strip()

    metadata: Dict[str, Any] = {}
    if extra_fields:
        metadata["extra_fields"] = extra_fields
    if source_review:
        metadata["source_review"] = source_review
    if source_review_title:
        metadata["source_review_title"] = source_review_title
    if source_doc_id:
        metadata["source_doc_id"] = source_doc_id
    return metadata


def _page_signals_open_access(page_text: str) -> bool:
    text = str(page_text or "").lower()
    markers = (
        "open access article",
        "this article is an open access article",
        "creative commons attribution",
        "distributed under the terms and conditions of the creative commons",
        "cc by license",
        "open access",
    )
    return any(marker in text for marker in markers)


def _extract_issn(message: Dict[str, Any]) -> str:
    for key in ("ISSN", "issn-type"):
        value = message.get(key)
        if isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    issn = str(item.get("value") or "").strip()
                    if issn:
                        return issn
                else:
                    issn = str(item or "").strip()
                    if issn:
                        return issn
        issn = str(value or "").strip()
        if issn:
            return issn
    return ""


def _extract_crossref_year(message: Dict[str, Any]) -> str:
    for key in ("published-print", "published-online", "issued", "created"):
        section = message.get(key) or {}
        parts = section.get("date-parts") if isinstance(section, dict) else None
        if isinstance(parts, list) and parts and isinstance(parts[0], list) and parts[0]:
            value = str(parts[0][0]).strip()
            if _YEAR_RE.fullmatch(value):
                return value
    return ""


def _extract_message_volume_issue(message: Dict[str, Any], citation: Dict[str, Any]) -> tuple[str, str]:
    volume = str(message.get("volume") or citation.get("volume") or "").strip()
    issue = str(message.get("issue") or citation.get("issue") or "").strip()
    return volume, issue


def _extract_article_number(message: Dict[str, Any], citation: Dict[str, Any], doi: str) -> str:
    for key in ("article-number", "page"):
        value = str(message.get(key) or citation.get("pages") or "").strip()
        if value:
            first = re.split(r"[-,;\s]", value, maxsplit=1)[0].strip()
            if first:
                return first

    suffix = str(doi or "").split("/")[-1].strip().lower()
    match = re.search(r"(\d{2})(\d{2})(\d{3,5})$", suffix)
    if match:
        return str(int(match.group(3)))
    return ""


def _normalize_issue_part(raw: str) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    if text.isdigit():
        return str(int(text))
    return text


def _mdpi_pdf_url(article_url: str) -> str:
    base = str(article_url or "").rstrip("/")
    if not base:
        return ""
    return f"{base}/pdf"


def _extract_mdpi_path_parts(doi: str) -> tuple[str, str, str]:
    suffix = str(doi or "").split("/")[-1].strip().lower()
    match = re.search(r"[a-z]+(\d{2})(\d{2})(\d{3,5})$", suffix)
    if not match:
        return "", "", ""
    return str(int(match.group(1))), str(int(match.group(2))), str(int(match.group(3)))


def _retraction_info(message: Dict[str, Any]) -> Dict[str, Any]:
    updates = message.get("update-to") or []
    if not isinstance(updates, list):
        updates = [updates]
    for item in updates:
        if not isinstance(item, dict):
            continue
        label = str(item.get("label") or item.get("type") or "").strip().lower()
        if "retract" in label:
            return {
                "is_retracted": True,
                "warning": "CrossRef metadata indicates a retraction update.",
            }
    return {"is_retracted": False, "warning": ""}


class ResearchResolver:
    def __init__(
        self,
        *,
        options: Dict[str, Any],
        progress_cb: Optional[Callable[[float, str, Optional[str]], None]] = None,
        is_cancelled_cb: Optional[Callable[[], bool]] = None,
        session: Optional[requests.Session] = None,
        sleep_fn: Callable[[float], None] = time.sleep,
    ) -> None:
        self.options = dict(options or {})
        self.progress_cb = progress_cb
        self.is_cancelled_cb = is_cancelled_cb
        self.sleep_fn = sleep_fn
        self.session = session or requests.Session()
        polite_email = str(
            self.options.get("unpaywall_email")
            or self.options.get("crossref_mailto")
            or ""
        ).strip()
        ua = "CortexSuiteResearchResolve/1.0"
        if polite_email:
            ua = f"{ua} (mailto:{polite_email})"
        self.session.headers.update(
            {
                "User-Agent": ua,
                "Accept": "application/json",
            }
        )
        self._last_crossref_at = 0.0

    @staticmethod
    def _web_headers() -> Dict[str, str]:
        return {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }

    @staticmethod
    def _pdf_probe_headers() -> Dict[str, str]:
        return {
            "Accept": "application/pdf,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }

    def resolve_all(self, citations: List[Dict[str, Any]]) -> Dict[str, Any]:
        total = len(citations)
        resolved: List[Dict[str, Any]] = []
        unresolved: List[Dict[str, Any]] = []

        if self.progress_cb:
            self.progress_cb(5, f"Resolving {total} citations", "research_resolve_start")

        for idx, citation in enumerate(citations, start=1):
            self._check_cancel()
            row_id = citation.get("row_id", idx)
            self._emit_progress(idx - 1, total, f"Resolving row {row_id}/{total}", "research_resolve_lookup")
            result = self.resolve_one(citation)
            if result.get("status") == "resolved":
                resolved.append(result["payload"])
            else:
                unresolved.append(result["payload"])
            self._emit_progress(idx, total, f"Resolved row {row_id}/{total}", "research_resolve_lookup")

        stats = {
            "total": total,
            "resolved_high": sum(1 for item in resolved if item.get("confidence") == "high"),
            "resolved_low": sum(1 for item in resolved if item.get("confidence") == "low"),
            "unresolved": len(unresolved),
            "open_access": sum(1 for item in resolved if (item.get("open_access") or {}).get("is_oa") is True),
            "closed_access": sum(1 for item in resolved if (item.get("open_access") or {}).get("is_oa") is False),
        }

        if self.progress_cb:
            self.progress_cb(100, "Research resolve complete", "done")

        return {
            "status": "completed",
            "resolved": resolved,
            "unresolved": unresolved,
            "stats": stats,
        }

    def resolve_one(self, citation: Dict[str, Any]) -> Dict[str, Any]:
        doi = _normalize_doi(citation.get("doi") or "")
        accession = str(citation.get("accession") or "").strip()

        if doi:
            resolved = self._resolve_by_doi(doi, citation)
            if resolved:
                return {"status": "resolved", "payload": resolved}

        if accession:
            resolved = self._resolve_by_accession(accession, citation)
            if resolved:
                return {"status": "resolved", "payload": resolved}

        searched = self._resolve_by_search(citation)
        if searched:
            return {"status": "resolved", "payload": searched}

        unresolved = {
            "row_id": citation.get("row_id"),
            "input_title": str(citation.get("title") or "").strip(),
            "reason": "No CrossRef match found",
            "best_candidates": self._best_candidates_for_unresolved(citation),
        }
        unresolved.update(_citation_source_metadata(citation))
        return {"status": "unresolved", "payload": unresolved}

    def _check_cancel(self) -> None:
        if self.is_cancelled_cb and self.is_cancelled_cb():
            raise RuntimeError("Cancelled by operator")

    def _emit_progress(self, done: int, total: int, message: str, stage: str) -> None:
        if not self.progress_cb:
            return
        frac = 0.0 if total <= 0 else min(1.0, max(0.0, done / float(total)))
        pct = 10 + int(frac * 85)
        self.progress_cb(pct, message, stage)

    def _respect_crossref_delay(self) -> None:
        elapsed = time.monotonic() - self._last_crossref_at
        delay = max(0.0, 0.5 - elapsed)
        if delay > 0:
            self.sleep_fn(delay)

    def _request_json(
        self,
        url: str,
        *,
        params: Optional[Dict[str, Any]] = None,
        source: str,
    ) -> Optional[Dict[str, Any]]:
        retryable_statuses = {429, 500, 502, 503, 504}
        for attempt, delay in enumerate((0.0, 2.0, 4.0, 8.0), start=1):
            self._check_cancel()
            if source == "crossref":
                self._respect_crossref_delay()
            try:
                response = self.session.get(url, params=params, timeout=30)
                if source == "crossref":
                    self._last_crossref_at = time.monotonic()
            except requests.RequestException as exc:
                if attempt >= 4:
                    logger.warning("%s request failed for %s: %s", source, url, exc)
                    return None
                self.sleep_fn(delay)
                continue

            if response.status_code == 404:
                return None
            if response.status_code in retryable_statuses:
                if attempt >= 4:
                    logger.warning("%s request failed for %s with status %s", source, url, response.status_code)
                    return None
                self.sleep_fn(delay)
                continue
            if response.status_code >= 400:
                logger.warning("%s request failed for %s with status %s", source, url, response.status_code)
                return None
            try:
                return response.json()
            except Exception as exc:
                logger.warning("%s JSON parse failed for %s: %s", source, url, exc)
                return None
        return None

    def _crossref_works(self, *, doi: str = "", params: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
        if doi:
            encoded = quote(doi, safe="")
            payload = self._request_json(f"https://api.crossref.org/works/{encoded}", source="crossref")
            if isinstance(payload, dict):
                return payload.get("message") if isinstance(payload.get("message"), dict) else None
            return None

        payload = self._request_json("https://api.crossref.org/works", params=params, source="crossref")
        if not isinstance(payload, dict):
            return None
        message = payload.get("message")
        if not isinstance(message, dict):
            return None
        items = message.get("items")
        if isinstance(items, list):
            return {"items": items}
        return None

    def _resolve_by_doi(self, doi: str, citation: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        message = self._crossref_works(doi=doi)
        if not message:
            return None
        return self._build_resolved_payload(
            citation=citation,
            message=message,
            source_api="crossref_doi",
            resolution_method="doi_direct",
            confidence="high",
            similarity=1.0,
        )

    def _resolve_by_accession(self, accession: str, citation: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        if not _PMID_RE.fullmatch(accession):
            return None

        crossref = self._crossref_works(params={"filter": f"alternative-id:{accession}", "rows": 5})
        items = crossref.get("items") if crossref else []
        if items:
            candidate = self._score_candidates(
                citation,
                items,
                source_api="crossref_alt_id",
                resolution_method="accession_lookup",
            )
            if candidate and candidate.get("confidence") == "high":
                return candidate

        ncbi = self._request_json(
            "https://api.ncbi.nlm.nih.gov/lit/ctxp/v1/pubmed/",
            params={"format": "csl", "id": accession},
            source="ncbi",
        )
        doi = _normalize_doi(str((ncbi or {}).get("DOI") or ""))
        if not doi:
            return None
        message = self._crossref_works(doi=doi)
        if not message:
            return None
        return self._build_resolved_payload(
            citation=citation,
            message=message,
            source_api="ncbi_pubmed",
            resolution_method="accession_lookup",
            confidence="high",
            similarity=1.0,
        )

    def _resolve_by_search(self, citation: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        first_author = _extract_first_author(citation.get("authors") or "")
        params = {
            "query.title": citation.get("title") or "",
            "rows": 5,
        }
        if first_author:
            params["query.author"] = first_author

        crossref = self._crossref_works(params=params)
        items = crossref.get("items") if crossref else []
        candidate = self._score_candidates(
            citation,
            items,
            source_api="crossref_title_author",
            resolution_method="title_author_search",
        )
        if candidate:
            return candidate

        broader_params: Dict[str, Any] = {"query": citation.get("title") or "", "rows": 5}
        year = _extract_year(citation.get("year"))
        if year:
            broader_params["filter"] = f"from-pub-date:{year}-01-01,until-pub-date:{year}-12-31"
        broader = self._crossref_works(params=broader_params)
        broader_items = broader.get("items") if broader else []
        return self._score_candidates(
            citation,
            broader_items,
            source_api="crossref_broad",
            resolution_method="broader_search",
        )

    def _best_candidates_for_unresolved(self, citation: Dict[str, Any]) -> List[Dict[str, Any]]:
        params: Dict[str, Any] = {"query": citation.get("title") or "", "rows": 3}
        fallback = self._crossref_works(params=params)
        items = fallback.get("items") if fallback else []
        return self._candidate_summaries(citation, items)[:3]

    def _candidate_summaries(self, citation: Dict[str, Any], items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        summaries = []
        target_title = str(citation.get("title") or "").strip()
        target_year = _extract_year(citation.get("year"))
        for item in items or []:
            title = self._message_title(item)
            similarity = _title_similarity(target_title, title)
            year_match = target_year and _extract_crossref_year(item) == target_year
            if year_match:
                similarity = min(1.0, similarity + 0.02)
            summaries.append(
                {
                    "title": title,
                    "doi": _normalize_doi(str(item.get("DOI") or "")),
                    "similarity": round(similarity, 3),
                }
            )
        return sorted(summaries, key=lambda item: item.get("similarity", 0.0), reverse=True)

    def _score_candidates(
        self,
        citation: Dict[str, Any],
        items: List[Dict[str, Any]],
        *,
        source_api: str,
        resolution_method: str,
    ) -> Optional[Dict[str, Any]]:
        if not items:
            return None

        target_year = _extract_year(citation.get("year"))
        best_item: Optional[Dict[str, Any]] = None
        best_similarity = 0.0

        for item in items:
            title = self._message_title(item)
            similarity = _title_similarity(citation.get("title") or "", title)
            item_year = _extract_crossref_year(item)
            if target_year and item_year and item_year == target_year:
                similarity = min(1.0, similarity + 0.03)
            if similarity > best_similarity:
                best_similarity = similarity
                best_item = item

        if not best_item or best_similarity < 0.7:
            return None

        confidence = "low"
        year_match = target_year and _extract_crossref_year(best_item) == target_year
        if best_similarity > 0.9 and year_match:
            confidence = "high"

        return self._build_resolved_payload(
            citation=citation,
            message=best_item,
            source_api=source_api,
            resolution_method=resolution_method,
            confidence=confidence,
            similarity=best_similarity,
        )

    def _message_title(self, message: Dict[str, Any]) -> str:
        title = message.get("title")
        if isinstance(title, list):
            return str(title[0] or "").strip() if title else ""
        return str(title or "").strip()

    def _merge_open_access(self, primary: Dict[str, Any], fallback: Dict[str, Any]) -> Dict[str, Any]:
        merged = dict(primary or _empty_open_access())
        extra = dict(fallback or _empty_open_access())
        if extra.get("is_oa") is True and merged.get("is_oa") is not True:
            merged["is_oa"] = True
            merged["oa_status"] = str(extra.get("oa_status") or "publisher_free_pdf").strip() or "publisher_free_pdf"
        elif merged.get("is_oa") is None and extra.get("is_oa") is False:
            merged["is_oa"] = False
            merged["oa_status"] = str(extra.get("oa_status") or "closed").strip() or "closed"

        if not str(merged.get("pdf_url") or "").strip() and str(extra.get("pdf_url") or "").strip():
            merged["pdf_url"] = str(extra.get("pdf_url") or "").strip()

        primary_source = str(merged.get("oa_source") or "").strip()
        fallback_source = str(extra.get("oa_source") or "").strip()
        if fallback_source:
            if primary_source:
                parts = [part for part in primary_source.split("+") if part]
                if fallback_source not in parts:
                    parts.append(fallback_source)
                merged["oa_source"] = "+".join(parts)
            else:
                merged["oa_source"] = fallback_source
        return merged

    def _probe_pdf_url(self, url: str) -> str:
        candidate = str(url or "").strip()
        if not candidate:
            return ""
        try:
            response = self.session.get(
                candidate,
                timeout=20,
                allow_redirects=True,
                stream=True,
                headers=self._pdf_probe_headers(),
            )
        except requests.RequestException:
            return ""
        try:
            if response.status_code >= 400:
                return ""
            content_type = str(response.headers.get("Content-Type") or "").lower()
            if "application/pdf" in content_type or response.url.lower().endswith(".pdf"):
                return str(response.url or candidate).strip()
            return ""
        finally:
            response.close()

    def _extract_pdf_candidates(self, html: str, base_url: str) -> List[str]:
        soup = BeautifulSoup(html or "", "html.parser")
        candidates: List[str] = []

        meta_pdf = soup.find("meta", attrs={"name": "citation_pdf_url"})
        if meta_pdf and meta_pdf.get("content"):
            candidates.append(urljoin(base_url, str(meta_pdf.get("content") or "").strip()))

        for link in soup.find_all("link"):
            href = str(link.get("href") or "").strip()
            link_type = str(link.get("type") or "").lower()
            if not href:
                continue
            if "pdf" in link_type or href.lower().endswith(".pdf"):
                candidates.append(urljoin(base_url, href))

        for anchor in soup.find_all("a"):
            href = str(anchor.get("href") or "").strip()
            if not href:
                continue
            href_low = href.lower()
            text_low = str(anchor.get_text(" ", strip=True) or "").lower()
            if href_low.endswith(".pdf") or "/pdf" in href_low or ("download" in text_low and "pdf" in text_low):
                candidates.append(urljoin(base_url, href))

        return list(dict.fromkeys(candidates))

    def _publisher_page_open_access(self, resolved_url: str) -> Dict[str, Any]:
        if not str(resolved_url or "").strip():
            return _empty_open_access()

        direct_pdf = self._probe_pdf_url(resolved_url)
        if direct_pdf:
            return {
                "is_oa": True,
                "oa_status": "publisher_free_pdf",
                "pdf_url": direct_pdf,
                "oa_source": "publisher_page",
            }

        try:
            response = self.session.get(
                str(resolved_url).strip(),
                timeout=20,
                allow_redirects=True,
                headers=self._web_headers(),
            )
        except requests.RequestException:
            return _empty_open_access()

        if response.status_code >= 400:
            return _empty_open_access()

        content_type = str(response.headers.get("Content-Type") or "").lower()
        final_url = str(response.url or resolved_url).strip()
        if "application/pdf" in content_type:
            return {
                "is_oa": True,
                "oa_status": "publisher_free_pdf",
                "pdf_url": final_url,
                "oa_source": "publisher_page",
            }
        if "html" not in content_type and "<html" not in response.text[:500].lower():
            return _empty_open_access()

        candidate_urls = self._extract_pdf_candidates(response.text, final_url)
        for candidate in candidate_urls:
            verified = self._probe_pdf_url(candidate)
            if verified:
                return {
                    "is_oa": True,
                    "oa_status": "publisher_free_pdf",
                    "pdf_url": verified,
                    "oa_source": "publisher_page",
                }
        if candidate_urls and _page_signals_open_access(response.text):
            return {
                "is_oa": True,
                "oa_status": "publisher_page_open_access",
                "pdf_url": candidate_urls[0],
                "oa_source": "publisher_page",
            }
        return _empty_open_access()

    def _publisher_landing_candidates(self, *, message: Dict[str, Any], citation: Dict[str, Any], doi: str, resolved_url: str) -> List[str]:
        candidates: List[str] = []
        if str(resolved_url or "").strip():
            candidates.append(str(resolved_url).strip())

        publisher = str(message.get("publisher") or "").strip().lower()
        issn = _extract_issn(message)
        journal_name = ""
        container = message.get("container-title")
        if isinstance(container, list) and container:
            journal_name = str(container[0] or "").strip()
        if not journal_name:
            journal_name = str(citation.get("journal") or "").strip()
        if "mdpi" in publisher and not issn and journal_name:
            ranking = classify_journal_authority(journal_name, journal_name)
            issn = str(ranking.get("journal_issn") or "").strip()
        volume, issue = _extract_message_volume_issue(message, citation)
        article_number = _extract_article_number(message, citation, doi)
        mdpi_volume, mdpi_issue, mdpi_article = _extract_mdpi_path_parts(doi)

        if "mdpi" in publisher and issn:
            if not volume:
                volume = mdpi_volume
            if not issue:
                issue = mdpi_issue
            if not article_number:
                article_number = mdpi_article
        if "mdpi" in publisher and issn and volume and issue and article_number:
            candidates.append(
                f"https://www.mdpi.com/{issn}/{volume}/{_normalize_issue_part(issue)}/{article_number}"
            )

        return list(dict.fromkeys(candidate for candidate in candidates if candidate))

    def _publisher_policy_open_access(self, *, message: Dict[str, Any], citation: Dict[str, Any], doi: str, landing_candidates: List[str]) -> Dict[str, Any]:
        publisher = str(message.get("publisher") or "").strip().lower()
        if "mdpi" in publisher and str(doi or "").lower().startswith("10.3390/"):
            mdpi_candidate = next((item for item in landing_candidates if "mdpi.com/" in item), "")
            if mdpi_candidate:
                return {
                    "is_oa": True,
                    "oa_status": "publisher_policy_open_access",
                    "pdf_url": _mdpi_pdf_url(mdpi_candidate),
                    "oa_source": "publisher_policy",
                }
        return _empty_open_access()

    def _open_access_info(self, doi: str, resolved_url: str, message: Dict[str, Any], citation: Dict[str, Any]) -> Dict[str, Any]:
        if not self.options.get("check_open_access", True):
            return _empty_open_access()
        email = str(self.options.get("unpaywall_email") or "").strip()
        result = _empty_open_access()
        landing_candidates = self._publisher_landing_candidates(
            message=message,
            citation=citation,
            doi=doi,
            resolved_url=resolved_url,
        )
        if not doi or not email:
            for candidate in landing_candidates:
                result = self._merge_open_access(result, self._publisher_page_open_access(candidate))
                if result.get("is_oa") is True:
                    return result
            final_result = self._merge_open_access(
                result,
                self._publisher_policy_open_access(
                    message=message,
                    citation=citation,
                    doi=doi,
                    landing_candidates=landing_candidates,
                ),
            )
            return final_result

        payload = self._request_json(
            f"https://api.unpaywall.org/v2/{quote(doi, safe='')}",
            params={"email": email},
            source="unpaywall",
        )
        if isinstance(payload, dict):
            location = payload.get("best_oa_location") or {}
            pdf_url = ""
            if isinstance(location, dict):
                pdf_url = str(location.get("url_for_pdf") or location.get("url") or "").strip()
            result = {
                "is_oa": payload.get("is_oa") if isinstance(payload.get("is_oa"), bool) else None,
                "oa_status": str(payload.get("oa_status") or "unknown").strip().lower() or "unknown",
                "pdf_url": pdf_url,
                "oa_source": "unpaywall",
            }
        if result.get("is_oa") is True and str(result.get("pdf_url") or "").strip():
            return result
        for candidate in landing_candidates:
            result = self._merge_open_access(result, self._publisher_page_open_access(candidate))
            if result.get("is_oa") is True:
                return result
        final_result = self._merge_open_access(
            result,
            self._publisher_policy_open_access(
                message=message,
                citation=citation,
                doi=doi,
                landing_candidates=landing_candidates,
            ),
        )
        return final_result

    def _journal_info(self, citation: Dict[str, Any], message: Dict[str, Any]) -> Dict[str, Any]:
        journal_name = ""
        container = message.get("container-title")
        if isinstance(container, list) and container:
            journal_name = str(container[0] or "").strip()
        if not journal_name:
            journal_name = str(citation.get("journal") or "").strip()
        issn = _extract_issn(message)

        ranking = {
            "journal_title": journal_name,
            "journal_issn": issn,
            "journal_quartile": "",
            "journal_sjr": 0.0,
            "journal_rank_global": 0,
        }
        if self.options.get("enrich_sjr", True):
            ranking = classify_journal_authority(journal_name, f"{issn}\n{citation.get('journal') or ''}")
            if not ranking.get("journal_title") and journal_name:
                ranking["journal_title"] = journal_name
            if not ranking.get("journal_issn") and issn:
                ranking["journal_issn"] = issn

        return {
            "name": str(ranking.get("journal_title") or journal_name or "").strip(),
            "issn": str(ranking.get("journal_issn") or issn or "").strip(),
            "sjr_quartile": str(ranking.get("journal_quartile") or "").strip(),
            "sjr_score": float(ranking.get("journal_sjr") or 0.0),
            "sjr_rank": int(ranking.get("journal_rank_global") or 0),
        }

    def _credibility_tier(self, message: Dict[str, Any], resolved_url: str, journal_name: str) -> str:
        combined = " ".join(
            [
                resolved_url,
                str(message.get("publisher") or ""),
                str(message.get("subtype") or ""),
                str(message.get("type") or ""),
                journal_name,
            ]
        ).lower()
        if any(marker in combined for marker in _PREPRINT_MARKERS):
            return "pre-print"
        if str(message.get("type") or "").strip().lower() in {"journal-article", "proceedings-article"}:
            return "peer-reviewed"
        tier_value, tier_key, _, _ = classify_credibility_tier_with_reason(
            text=f"{resolved_url}\n{journal_name}\n{message.get('publisher') or ''}",
            source_type="Other",
            availability_status="available",
        )
        return tier_key if tier_value >= 0 else "unclassified"

    def _build_resolved_payload(
        self,
        *,
        citation: Dict[str, Any],
        message: Dict[str, Any],
        source_api: str,
        resolution_method: str,
        confidence: str,
        similarity: float,
    ) -> Dict[str, Any]:
        doi = _normalize_doi(str(message.get("DOI") or citation.get("doi") or ""))
        resolved_url = _pick_primary_url(message, doi)
        open_access = self._open_access_info(doi, resolved_url, message, citation)
        journal = self._journal_info(citation, message)
        retraction = _retraction_info(message)

        payload = {
            "row_id": citation.get("row_id"),
            "input_title": str(citation.get("title") or "").strip(),
            "resolved_doi": doi,
            "resolved_url": resolved_url,
            "source_api": source_api,
            "confidence": confidence,
            "open_access": open_access,
            "journal": journal,
            "credibility_tier": self._credibility_tier(message, resolved_url, journal.get("name") or ""),
            "resolution_method": resolution_method,
            "publisher": str(message.get("publisher") or "").strip(),
            "type": str(message.get("type") or "").strip(),
            "matched_title": self._message_title(message),
            "title_similarity": round(similarity, 3),
            "retraction": retraction,
        }
        payload.update(_citation_source_metadata(citation))
        return payload


def run_research_resolve(
    *,
    payload: Dict[str, Any],
    run_dir: Path,
    progress_cb: Optional[Callable[[float, str, Optional[str]], None]] = None,
    is_cancelled_cb: Optional[Callable[[], bool]] = None,
) -> Dict[str, Any]:
    resolver = ResearchResolver(
        options=dict(payload.get("options") or {}),
        progress_cb=progress_cb,
        is_cancelled_cb=is_cancelled_cb,
    )
    output = resolver.resolve_all(list(payload.get("citations") or []))
    out_path = Path(run_dir) / "research_resolve_result.json"
    out_path.write_text(json.dumps(output, ensure_ascii=True, indent=2), encoding="utf-8")
    return output
